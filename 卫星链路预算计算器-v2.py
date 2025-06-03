"""
卫星链路预算计算器

功能：
1. 支持输入链路预算参数，包括频率、距离、高度、极化方式等。
2. 计算链路预算结果，包括信号强度、噪声功率、信噪比、链路长度等。
3. 支持导出链路预算结果到Excel文件。
依赖库 pip install customtkinter openpyxl pandas
打包命令：pyinstaller --onedir --windowed  --hidden-import customtkinter  --hidden-import openpyxl --collect-all customtkinter  "D:\PySimpleGUI.py"
"""
import tkinter as tk
import customtkinter as ctk
from openpyxl import Workbook
from openpyxl.styles import Font, Border, Side, Alignment  # Add this import
from tkinter import filedialog, messagebox
from datetime import datetime
import math
from LinkCalculator import LinkCalculator, UnitConverter
from SafeMath import safe_eval, format_result
from IOHandler import InputHandler, ResultDisplay
from parameters import PARAM_MAPPING, PARAM_GROUPS, RESULT_CATEGORIES


ctk.set_appearance_mode("System")  # 跟随系统主题
ctk.set_default_color_theme("blue")  # 使用官方主题配色

# 主应用类
# 常量定义
DEFAULT_FONT = ("微软雅黑", 12)
TITLE_FONT = ("微软雅黑", 14, "bold")
GROUP_TITLE_FONT = ("微软雅黑", 12, "bold")
GROUP_TITLE_COLOR = "#165DFF"
STATUS_BAR_FONT = ("微软雅黑", 10)
STATUS_BAR_BG_COLOR = "#f0f0f0"
STATUS_BAR_TEXT_COLOR = "#333"

class SatelliteLinkBudgetCalculator:
    def __init__(self, root):
        self.root = root
        self.root.title("卫星链路预算计算器")
        self.root.geometry("1024x768")  # 设置默认窗口尺寸
        self.gt_label = None  # 用于显示G/T值的标签
        self._init_ui()
        

    def _init_ui(self):
        """初始化用户界面"""
        self.main_frame = ctk.CTkFrame(self.root)
        self.main_frame.pack(fill=tk.BOTH, expand=True, padx=10, pady=10)

        self._create_link_type_selector()
        self._create_toolbar()
        self._create_content_frames()
        self._create_status_bar()
        self.change_link_type(self.link_type_var.get())

    def _create_link_type_selector(self):
        """创建链路类型选择器"""
        self.link_type_var = tk.StringVar(value="星-地下行")
        link_type_label = ctk.CTkLabel(self.main_frame, text="选择链路类型:", font=DEFAULT_FONT)
        link_type_label.pack(pady=(5, 0))
        link_type_optionmenu = ctk.CTkOptionMenu(
            self.main_frame,
            variable=self.link_type_var,
            values=["星-地下行", "星-地上行", "地-地下行", "地-地上行"],
            command=self.change_link_type
        )
        link_type_optionmenu.pack(pady=(0, 5))

    def _setup_gt_display(self, link_type):
        """设置动态G/T值显示框"""
        # 定义不同链路类型的显示配置
        gt_config = {
            "星-地上行": "卫星G/T值 (dB/K):",
            "星-地下行": "终端G/T值 (dB/K):",
            "地-地上行": "基站G/T值 (dB/K):",
            "地-地下行": "终端G/T值 (dB/K):"
        }
        
        # 清除旧标签
        if hasattr(self, 'gt_label') and self.gt_label:
            self.gt_label.destroy()
        
        # 创建新标签组件
        gt_frame = self.input_handler.gt_frame
        ctk.CTkLabel(gt_frame, 
                    text=gt_config.get(link_type, "G/T值:"), 
                    width=180, 
                    anchor="w").pack(side=tk.LEFT, padx=5)
        
        self.gt_label = ctk.CTkLabel(gt_frame, 
                                   text="0.00", 
                                   width=120, 
                                   anchor="e")
        self.gt_label.pack(side=tk.RIGHT, padx=5)


    def _create_toolbar(self):
        """创建工具栏"""
        self.toolbar = ctk.CTkFrame(self.main_frame)
        self.toolbar.pack(fill=tk.X, pady=5)

        self._create_left_toolbar_buttons()
        self._create_right_toolbar_buttons()

    def _create_left_toolbar_buttons(self):
        """创建左侧工具栏按钮"""
        self.button_frame = ctk.CTkFrame(self.toolbar, fg_color="transparent")
        self.button_frame.pack(side=tk.LEFT, padx=5)

        buttons = [
            ("计算", self.calculate, "#165DFF", 100),
            ("重置", self.reset, "#6B7280", 100),
            ("输出计算报告", self.generate_report, "#36B37E", 120),
            ("详细计算公式", self.show_detailed_calculation, "#36B37E", 120),  # 新增按钮
            ("单位转换器", self.show_unit_converter, "#FFA500", 120),  # 新增按钮
        ]
        for text, command, color, width in buttons:
            button = ctk.CTkButton(
                self.button_frame, text=text, command=command,
                font=DEFAULT_FONT, fg_color=color, width=width
            )
            button.pack(side=tk.LEFT, padx=5)

    def _create_right_toolbar_buttons(self):
        """创建右侧工具栏按钮"""
        self.right_button_frame = ctk.CTkFrame(self.toolbar, fg_color="transparent")
        self.right_button_frame.pack(side=tk.RIGHT, padx=5)

        self.theme_switch = ctk.CTkSwitch(
            self.right_button_frame, text="深色模式", command=self.toggle_theme,
            font=STATUS_BAR_FONT
        )
        self.theme_switch.pack(side=tk.LEFT, padx=10)

    def _create_content_frames(self):
        """创建内容框架"""
        self.content_frame = ctk.CTkFrame(self.main_frame)
        self.content_frame.pack(fill=tk.BOTH, expand=True, pady=10)

        self.input_frame = ctk.CTkFrame(self.content_frame, corner_radius=10)
        self.input_frame.pack(side=tk.LEFT, fill=tk.BOTH, expand=True, padx=(0, 5))

        self.result_frame = ctk.CTkFrame(self.content_frame, corner_radius=10)
        self.result_frame.pack(side=tk.RIGHT, fill=tk.BOTH, expand=True, padx=(5, 0))

        self.result_display = ResultDisplay()
        self.result_display.create_result_display(self.result_frame)

    def _create_status_bar(self):
        """创建状态栏"""
        self.status_var = tk.StringVar(value="支持公式化输入，例如：sin(30)、arctan(1)、53-10*log(16)")
        self.status_bar = ctk.CTkLabel(
            self.main_frame, textvariable=self.status_var,
            font=STATUS_BAR_FONT, fg_color=STATUS_BAR_BG_COLOR, text_color=STATUS_BAR_TEXT_COLOR
        )
        self.status_bar.pack(fill=tk.X, pady=5)

    def change_link_type(self, link_type):
        """切换链路类型时更新输入处理器"""
        self._clear_input_frame()
        params = self._setup_link_defaults(link_type)
        
        # 创建新的输入处理器时传递当前链路类型
        self.input_handler = InputHandler(
            self.input_frame, 
            params,
            link_type  # 传递当前选择的链路类型
        )
        
        self._setup_input_handler(params)
        self._setup_gt_display(link_type)

    def _clear_input_frame(self):
        """清除输入框"""
        for widget in self.input_frame.winfo_children():
            widget.destroy()

    def _setup_link_defaults(self, link_type):
        """从PARAM_GROUPS智能组合链路参数"""
        config = PARAM_GROUPS[link_type]
        base_config = PARAM_GROUPS[config["base"]]
        
        def get_param(param):
            value = PARAM_MAPPING[param]["default_value"]
            return value.get(link_type, value) if isinstance(value, dict) else value

        # 合并基础参数
        merged_params = {
            param: get_param(param)
            for group in ["common", "optional", "beam_params", "interference_params"]
            for param in base_config[group]
        }
        
        # 添加收发端参数
        merged_params.update({
            param: get_param(param)
            for param in config["tx_params"] + config["rx_params"]
        })
        
        return merged_params

    def _setup_input_handler(self, params):
        """设置输入处理器"""
        self.input_handler = InputHandler(
            self.input_frame, 
            params,
            self.link_type_var.get()  # 直接传递当前链路类型
        )
        self.input_handler.create_input_form(self.input_frame)


    def _get_input_params(self):
        """获取输入参数（完整版）"""
        # 主要作用：
        # ① 收集用户实际输入的数值
        # ② 处理参数间的依赖关系
        # ③ 返回用于链路计算的有效参数集合
        self.input_handler.trigger_all_focus_out()
        link_type = self.link_type_var.get()
        
        # 获取参数配置
        config = PARAM_GROUPS[link_type]
        base_config = PARAM_GROUPS[config["base"]]
        
        # 合并参数组
        all_params = {
            "common": base_config["common"],
            "optional": base_config["optional"],
            "beam_params": base_config["beam_params"],
            "interference_params": base_config["interference_params"],
            "tx": config["tx_params"],
            "rx": config["rx_params"]
        }
        # 基础参数
        input_params = {
            param: self.input_handler.get_numeric_value(param) 
            for param in all_params["common"]
        }
        # 可选参数
        for param in all_params["optional"]+all_params["beam_params"]:
            input_params[param] = self.input_handler.get_numeric_value(param) if self.input_handler.flags[param].get() else 0

        # 干扰参数特殊处理
        input_params["interference_psd"] =   self.input_handler.get_numeric_value("interference_psd") if self.input_handler.flags["interference_psd"].get() else -math.inf  # 新增干扰参数

        # 发射端参数
        tx_param = all_params["tx"][0]
        input_params["tx_eirp"] = self.input_handler.get_numeric_value(tx_param)

        # 接收端参数
        rx_ant, rx_nf, rx_nt = all_params["rx"]
        input_params.update({
            "rx_antenna_gain": self.input_handler.get_numeric_value(rx_ant),
            "rx_noise_figure": self.input_handler.get_numeric_value(rx_nf),
            "rx_noise_temp": self.input_handler.get_numeric_value(rx_nt)
        })

        if link_type in ["地-地上行", "地-地下行"]:
            # 合并进地面信道状态信息 
            input_params.update(self.input_handler.get_terrestrial_link_parameters()) 
            print(input_params)

        return input_params

    def calculate(self):
        try:
            self.status_var.set("正在计算...")
            input_params = self._get_input_params()

            calculator = LinkCalculator()
            link_type = self.link_type_var.get()
            # 执行计算
            self.results_temp = calculator.perform_calculations(input_params, link_type)  # 将结果保存为类属性

            # 更新G/T值显示
            if self.gt_label and "gt_ratio" in self.results_temp:
                self.gt_label.configure(text=format_result(self.results_temp["gt_ratio"]))
            """
            if self.link_type_var.get() == "星-地上行": # 星-地上行链路
                results["链路性能"].append(("卫星G/T值", self.results_temp["gt_ratio"], "dB/K"))
            elif self.link_type_var.get() in ["星-地下行", "地-地下行"]: # 下行链路
                results["链路性能"].append(("终端G/T值", self.results_temp["gt_ratio"], "dB/K"))
            elif self.link_type_var.get() =="地-地上行": # 地-地上行链路
                results["链路性能"].append(("基站G/T值", self.results_temp["gt_ratio"], "dB/K"))
            """
            self.result_display.update_results(self.results_temp, self.link_type_var.get())
            self.status_var.set("计算完成")
        except Exception as e:
            messagebox.showerror("计算错误", f"计算过程中出现错误: {str(e)}")
            self.status_var.set("计算失败，请检查输入")

    def reset(self):
        self.input_handler.reset_params()
        # 清空G/T值显示
        if hasattr(self, 'gt_label') and self.gt_label:
            self.gt_label.configure(text="0.00")
        # 强制清空结果区域
        self.result_display.clear_results()
        # 刷新界面
        self.result_frame.update_idletasks()
        self.status_var.set("所有参数和结果已重置")

    def generate_report(self):
        try:
            # 获取当前时间作为文件名的一部分
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            # 弹出文件保存对话框，让用户选择保存路径
            file_path = filedialog.asksaveasfilename(
                defaultextension=".xlsx",
                filetypes=[("Excel 文件", "*.xlsx")],
                initialfile=f"卫星链路预算仿真报告_{timestamp}.xlsx"
            )
            if not file_path:  # 如果用户取消选择，直接返回
                return

            # 创建一个新的 Excel 工作簿
            wb = Workbook()
            ws = wb.active
            ws.title = "仿真报告"

            # 设置表格样式
            header_font = Font(name='微软雅黑', size=12, bold=True)
            data_font = Font(name='微软雅黑', size=11)
            border = Border(left=Side(style='thin'), 
                         right=Side(style='thin'), 
                         top=Side(style='thin'), 
                         bottom=Side(style='thin'))
            alignment = Alignment(horizontal='center', vertical='center')

            # 定义参数标签 (英文参数名 -> 中文名称)
            param_labels = {
                param: (PARAM_MAPPING[param]["ch_name"], PARAM_MAPPING[param]["unit"])
                for param in PARAM_MAPPING
            }

            # 写入报告标题
            ws.merge_cells('A1:C1')
            title_cell = ws['A1']
            title_cell.value = "卫星链路预算仿真报告"
            title_cell.font = Font(name='微软雅黑', size=14, bold=True)
            title_cell.alignment = alignment
            title_cell.border = border

            ws.append([])  # 空行

            # 写入输入参数
            ws.append(["输入参数", "单位", "值"])
            ws['A'+str(ws.max_row)].font = header_font
            ws['B'+str(ws.max_row)].font = header_font
            ws['C'+str(ws.max_row)].font = header_font

            # 写入输入参数并应用样式
            for param, (label, unit) in param_labels.items():
                if param in self.input_handler.params:
                    flag_key = {
                        "atmospheric_loss": "atmospheric_loss",
                        "scintillation_loss": "scintillation_loss",
                        "polarization_loss": "polarization_loss",
                        "rain_rate": "rain_rate",
                        "beam_edge_loss": "beam_edge_loss",
                        "scan_loss": "scan_loss",
                        "link_margin": "link_margin",
                        "interference_psd": "interference_psd"
                    }.get(param, None)

                    # 如果参数没有标志位，或者标志位为 True，则显示
                    if not flag_key or self.input_handler.flags[flag_key].get():
                        value = self.input_handler.params[param].get()
                        try:
                            numeric_value = float(safe_eval(value)) if value else 0
                            ws.append([label, unit, numeric_value])
                        except:
                            ws.append([label, unit, 0])

            # 设置列宽
            ws.column_dimensions['A'].width = 30
            ws.column_dimensions['B'].width = 15
            ws.column_dimensions['C'].width = 20

            # 写入计算结果
            if hasattr(self, "results_temp"):
                ws.append([])  # 空行
                ws.append(["计算结果", "单位", "值"])
                ws['A'+str(ws.max_row)].font = header_font
                ws['B'+str(ws.max_row)].font = header_font
                ws['C'+str(ws.max_row)].font = header_font

                """
                # 根据链路类型添加G/T值
                if self.link_type_var.get() == "星-地上行":
                    result_items.append(("卫星G/T值", "dB/K"))
                elif self.link_type_var.get() == "星-地下行" or self.link_type_var.get() == "地-地下行":
                    result_items.append(("终端G/T值", "dB/K"))
                else:  # 地对地场景
                    result_items.append(("基站G/T值", "dB/K"))
                """
                # 根据链路类型选择结果分类
                link_category = "卫星链路" if "星" in self.link_type_var.get() else "地面链路"
                for category in RESULT_CATEGORIES[link_category].values():
                    for item in category:
                        key = item["key"]
                        ws.append([
                            item["label"],
                            item["unit"],
                            self.results_temp.get(key, 0)
                        ])            

            # 应用边框样式到所有单元格
            for row in ws.iter_rows():
                for cell in row:
                    cell.border = border

            # 保存 Excel 文件
            wb.save(file_path)
            messagebox.showinfo("报告生成成功", f"仿真报告已保存至:\n{file_path}")
        except Exception as e:
            messagebox.showerror("报告生成失败", f"生成仿真报告时出错:\n{str(e)}")


    def show_detailed_calculation(self):
        """显示详细计算步骤"""
        try:
            # 获取输入参数
            input_params = self._get_input_params()
            # 调用详细计算方法
            calculator = LinkCalculator()
            details = calculator.detailed_calculation(input_params)
            
            # 创建新窗口显示结果
            detail_window = ctk.CTkToplevel(self.root)
            detail_window.title("详细计算步骤")
            detail_window.geometry("800x600")

            # 将新窗口设置为主窗口的子窗口
            detail_window.transient(self.root)
            # 将新窗口提升到最前面
            detail_window.lift()

            # 创建文本框显示详细步骤
            text_box = ctk.CTkTextbox(detail_window, wrap="word")
            text_box.pack(fill="both", expand=True, padx=10, pady=10)
            
            # 格式化输出
            for detail in details:
                text_box.insert("end", f"步骤: {detail['步骤']}\n")
                text_box.insert("end", f"公式: {detail['公式']}\n")
                text_box.insert("end", f"参数: {detail['参数']}\n")
                text_box.insert("end", f"结果: {detail['结果']}\n\n")
                
        except Exception as e:
            messagebox.showerror("错误", f"无法显示详细计算步骤: {str(e)}")

    def toggle_theme(self):
        """ 深色模式 """
        ctk.set_appearance_mode("dark" if self.theme_switch.get() else "light")


    def show_unit_converter(self):
        """显示单位转换器窗口"""
        converter_window = ctk.CTkToplevel(self.root)
        converter_window.title("单位转换器")
        converter_window.geometry("600x400")
        
        # 主框架
        main_frame = ctk.CTkFrame(converter_window)
        main_frame.pack(fill="both", expand=True, padx=10, pady=10)

        # 创建转换器实例
        converter = UnitConverter()

        # 转换类型选择
        conversion_var = ctk.StringVar(value=list(converter.converters.keys())[0])
        conversion_label = ctk.CTkLabel(main_frame, text="选择转换类型:", font=DEFAULT_FONT)
        conversion_label.pack(pady=5)
        conversion_menu = ctk.CTkOptionMenu(
            main_frame,
            variable=conversion_var,
            values=list(converter.converters.keys())
        )
        conversion_menu.pack(pady=5)

        # 输入和输出框架
        io_frame = ctk.CTkFrame(main_frame, fg_color="transparent")
        io_frame.pack(fill="x", pady=10)

        input_var = ctk.StringVar()
        output_var = ctk.StringVar()

        last_focused = None

        def on_input_focus(event):
            nonlocal last_focused
            last_focused = "input"

        def on_output_focus(event):
            nonlocal last_focused
            last_focused = "output"

        def perform_conversion():
            input_value = ""
            direction = 0
            if last_focused == "input":
                try:
                    input_value = float(safe_eval(input_var.get(), sign_massagebox=False))
                except Exception:
                    output_var.set("输入无效")
                direction = 0
            elif last_focused == "output":
                try:
                    input_value = float(safe_eval(output_var.get(), sign_massagebox=False))
                except Exception:
                    input_var.set("输入无效")
                direction = 1
            else:
                return

            try:
                # 使用 safe_eval 解析公式
                numeric_value = input_value
                conversion_type = conversion_var.get()
                result = converter.convert(conversion_type, numeric_value, direction)
                if result is not None:
                    if direction == 0:
                        output_var.set(f"{result:.4f}")
                    else:
                        input_var.set(f"{result:.4f}")
                else:
                    if direction == 0:
                        output_var.set("输入无效")
                    else:
                        input_var.set("输入无效")
            except Exception as e:
                if direction == 0:
                    output_var.set("公式解析错误")
                else:
                    input_var.set("公式解析错误")

        input_unit_label = ctk.CTkLabel(io_frame, text="", font=DEFAULT_FONT)
        input_unit_label.pack(side=tk.LEFT, padx=5)
        input_entry = ctk.CTkEntry(io_frame, textvariable=input_var, width=150)
        input_entry.pack(side=tk.LEFT, padx=5)
        input_entry.bind("<FocusIn>", on_input_focus)
        input_entry.bind("<KeyRelease>", lambda event: perform_conversion())

        arrow_label = ctk.CTkLabel(io_frame, text="⇄", font=DEFAULT_FONT)
        arrow_label.pack(side=tk.LEFT, padx=10)

        output_unit_label = ctk.CTkLabel(io_frame, text="", font=DEFAULT_FONT)
        output_unit_label.pack(side=tk.LEFT, padx=5)
        output_entry = ctk.CTkEntry(io_frame, textvariable=output_var, width=150)
        output_entry.pack(side=tk.LEFT, padx=5)
        output_entry.bind("<FocusIn>", on_output_focus)
        output_entry.bind("<KeyRelease>", lambda event: perform_conversion())

        def update_units(*args):
            # 清空输入框数值
            input_var.set("")
            output_var.set("")

            conversion_type = conversion_var.get()
            input_unit_label.configure(text=converter.converters[conversion_type]["units"][0])
            output_unit_label.configure(text=converter.converters[conversion_type]["units"][1])

        conversion_var.trace_add("write", update_units)
        update_units()


if __name__ == "__main__":
    root = ctk.CTk()
    app = SatelliteLinkBudgetCalculator(root)
    root.mainloop()

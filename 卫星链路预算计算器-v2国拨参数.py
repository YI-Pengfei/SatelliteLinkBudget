"""
卫星链路预算计算器

功能：
1. 支持输入链路预算参数，包括频率、距离、高度、极化方式等。
2. 计算链路预算结果，包括信号强度、噪声功率、信噪比、链路长度等。
3. 支持导出链路预算结果到Excel文件。
依赖库 pip install customtkinter openpyxl pandas
打包命令：pyinstaller --onedir --windowed  --hidden-import customtkinter  --hidden-import openpyxl --collect-all customtkinter  "D:\PySimpleGUI.py"
"""
import tkinter as tk
import customtkinter as ctk
from openpyxl import Workbook
from openpyxl.styles import Font, Border, Side, Alignment  # Add this import
from tkinter import filedialog, messagebox
from datetime import datetime
import math
from LinkCalculator import LinkCalculator
from SafeMath import safe_eval, format_result
from IOHandler import InputHandler, ResultDisplay

ctk.set_appearance_mode("System")  # 跟随系统主题
ctk.set_default_color_theme("blue")  # 使用官方主题配色

# 主应用类
# 常量定义
DEFAULT_FONT = ("微软雅黑", 12)
TITLE_FONT = ("微软雅黑", 14, "bold")
GROUP_TITLE_FONT = ("微软雅黑", 12, "bold")
GROUP_TITLE_COLOR = "#165DFF"
STATUS_BAR_FONT = ("微软雅黑", 10)
STATUS_BAR_BG_COLOR = "#f0f0f0"
STATUS_BAR_TEXT_COLOR = "#333"

class SatelliteLinkBudgetCalculator:
    def __init__(self, root):
        self.root = root
        self.root.title("卫星链路预算计算器")
        self.root.geometry("1024x768")  # 设置默认窗口尺寸
        self._init_ui()

    def _init_ui(self):
        """初始化用户界面"""
        self.main_frame = ctk.CTkFrame(self.root)
        self.main_frame.pack(fill=tk.BOTH, expand=True, padx=10, pady=10)

        self._create_link_type_selector()
        self._init_gt_labels()
        self._create_toolbar()
        self._create_content_frames()
        self._create_status_bar()
        self.change_link_type(self.link_type_var.get())

    def _create_link_type_selector(self):
        """创建链路类型选择器"""
        self.link_type_var = tk.StringVar(value="星-地下行")
        link_type_label = ctk.CTkLabel(self.main_frame, text="选择链路类型:", font=DEFAULT_FONT)
        link_type_label.pack(pady=(5, 0))
        link_type_optionmenu = ctk.CTkOptionMenu(
            self.main_frame,
            variable=self.link_type_var,
            values=["星-地下行", "星-地上行", "地-地下行", "地-地上行"],
            command=self.change_link_type
        )
        link_type_optionmenu.pack(pady=(0, 5))

    def _init_gt_labels(self):
        """初始化G/T值显示框"""
        self.satellite_gt_label = None
        self.terminal_gt_label = None
        self.bs_gt_label = None

    def _create_toolbar(self):
        """创建工具栏"""
        self.toolbar = ctk.CTkFrame(self.main_frame)
        self.toolbar.pack(fill=tk.X, pady=5)

        self._create_left_toolbar_buttons()
        self._create_right_toolbar_buttons()

    def _create_left_toolbar_buttons(self):
        """创建左侧工具栏按钮"""
        self.button_frame = ctk.CTkFrame(self.toolbar, fg_color="transparent")
        self.button_frame.pack(side=tk.LEFT, padx=5)

        buttons = [
            ("计算", self.calculate, "#165DFF", 100),
            ("重置", self.reset, "#6B7280", 100),
            ("输出计算报告", self.generate_report, "#36B37E", 120),
            ("详细计算公式", self.show_detailed_calculation, "#36B37E", 120),  # 新增按钮
            # ("单位转换器", self.show_unit_converter, "#FFA500", 120),  # 新增按钮
        ]
        for text, command, color, width in buttons:
            button = ctk.CTkButton(
                self.button_frame, text=text, command=command,
                font=DEFAULT_FONT, fg_color=color, width=width
            )
            button.pack(side=tk.LEFT, padx=5)

    def _create_right_toolbar_buttons(self):
        """创建右侧工具栏按钮"""
        self.right_button_frame = ctk.CTkFrame(self.toolbar, fg_color="transparent")
        self.right_button_frame.pack(side=tk.RIGHT, padx=5)

        self.theme_switch = ctk.CTkSwitch(
            self.right_button_frame, text="深色模式", command=self.toggle_theme,
            font=STATUS_BAR_FONT
        )
        self.theme_switch.pack(side=tk.LEFT, padx=10)

    def _create_content_frames(self):
        """创建内容框架"""
        self.content_frame = ctk.CTkFrame(self.main_frame)
        self.content_frame.pack(fill=tk.BOTH, expand=True, pady=10)

        self.input_frame = ctk.CTkFrame(self.content_frame, corner_radius=10)
        self.input_frame.pack(side=tk.LEFT, fill=tk.BOTH, expand=True, padx=(0, 5))

        self.result_frame = ctk.CTkFrame(self.content_frame, corner_radius=10)
        self.result_frame.pack(side=tk.RIGHT, fill=tk.BOTH, expand=True, padx=(5, 0))

        self.result_display = ResultDisplay()
        self.result_display.create_result_display(self.result_frame)

    def _create_status_bar(self):
        """创建状态栏"""
        self.status_var = tk.StringVar(value="支持公式化输入，例如：53-10*log(16), 2*10**(3/10)")
        self.status_bar = ctk.CTkLabel(
            self.main_frame, textvariable=self.status_var,
            font=STATUS_BAR_FONT, fg_color=STATUS_BAR_BG_COLOR, text_color=STATUS_BAR_TEXT_COLOR
        )
        self.status_bar.pack(fill=tk.X, pady=5)

    def change_link_type(self, link_type):
        """切换链路类型"""
        self._clear_input_frame()
        params = self._get_link_params(link_type)
        self._setup_input_handler(params)
        self._setup_gt_display(link_type)

    def _clear_input_frame(self):
        """清除输入框"""
        for widget in self.input_frame.winfo_children():
            widget.destroy()

    def _get_link_params(self, link_type):
        """获取链路参数"""
        if link_type in ["星-地下行", "星-地上行"]:
            common_params = {
                "frequency": "1.71" if link_type == "星-地上行" else "1.81",
                "bandwidth": "5" if link_type == "星-地上行" else "5",
                "satellite_height": "515",
                "satellite_scan_angle": "0",
                "atmospheric_loss": "0.1",
                "scintillation_loss": "0.3",
                "polarization_loss": "3",
                "rain_rate": "50",
                "beam_edge_loss": "1",
                "scan_loss": "4",
                "link_margin": "3",
                "interference_psd": "-inf"
            }
            specific_params = {
                "星-地上行": {
                    "terminal_eirp": "23-30-3",
                    "satellite_antenna_gain": "35.2",
                    "satellite_noise_figure": "2.4",
                    "satellite_noise_temp": "290",
                },
                "星-地下行": {
                    "satellite_eirp": "48.2",
                    "terminal_noise_figure": "9",
                    "terminal_noise_temp": "290",
                    "terminal_antenna_gain": "-3",
                }
            }
            return {**common_params, **specific_params[link_type]}
        else: # 地对地场景的输入参数
            common_params = {
                "frequency": "1.71" if link_type == "地-地上行" else "1.81",
                "bandwidth": "5" if link_type == "地-地上行" else "5",
                "distance": "1", # 收发端2d距离 km
                "beam_edge_loss": "1",
                "interference_psd": "-inf"
            }
            specific_params = {
                "地-地上行": {
                    "terminal_eirp": "23-30-3",
                    "bs_antenna_gain": "30.72",
                    "bs_noise_figure": "2.4",
                    "bs_noise_temp": "290",
                },
                "地-地下行": {
                    "bs_eirp": "46+30+22.5",  # dBW 基站EIRP=发射功率(dBm)+天线增益(dBi)+30
                    "terminal_noise_figure": "7",
                    "terminal_noise_temp": "290",
                    "terminal_antenna_gain": "-3",
                }
            }
            return {**common_params, **specific_params[link_type]} 

    def _setup_input_handler(self, params):
        """设置输入处理器"""
        self.input_handler = InputHandler(self.input_frame, params)
        self.input_handler.create_input_form(self.input_frame)

    def _setup_gt_display(self, link_type):
        """设置G/T值显示框"""
        gt_frame = self.input_handler.gt_frame
        if link_type == "星-地上行":
            self._create_gt_label(gt_frame, "卫星G/T值 (dB/K):", "satellite_gt_label") # 创建卫星G/T值显示框
            self.terminal_gt_label = None  # 隐藏终端G/T值显示框
            self.bs_gt_label = None  # 隐藏基站G/T值显示框
        elif link_type == "地-地上行":
            self._create_gt_label(gt_frame, "基站G/T值 (dB/K):", "bs_gt_label") # 创建基站G/T值显示框
            self.terminal_gt_label = None  # 隐藏终端G/T值显示框
            self.satellite_gt_label = None  # 隐藏卫星G/T值显示框
        else: # 星-地下行、地-地下行
            self.satellite_gt_label = None  # 隐藏卫星G/T值显示框
            self._create_gt_label(gt_frame, "终端G/T值 (dB/K):", "terminal_gt_label")
            self.bs_gt_label = None  # 隐藏基站G/T值显示框

    def _create_gt_label(self, parent, text, attr_name):
        """创建G/T值标签"""
        ctk.CTkLabel(parent, text=text, width=180, anchor="w").pack(side=tk.LEFT, padx=5)
        label = ctk.CTkLabel(parent, text="0.00", width=120, anchor="e")
        label.pack(side=tk.RIGHT, padx=5)
        setattr(self, attr_name, label)

    def _get_input_params(self):
        """获取输入参数"""
        self.input_handler.trigger_all_focus_out()

        if self.link_type_var.get() == "星-地上行" or self.link_type_var.get() == "星-地下行":
            # 获取公共输入参数
            input_params = {
                "frequency": self.input_handler.get_numeric_value("frequency"),  # GHz,
                "bandwidth": self.input_handler.get_numeric_value("bandwidth"),  # MHz,
                "satellite_scan_angle": self.input_handler.get_numeric_value("satellite_scan_angle"),  # 度,
                "satellite_height": self.input_handler.get_numeric_value("satellite_height"),  # km,
                
                "atmospheric_loss": self.input_handler.get_numeric_value("atmospheric_loss") if self.input_handler.flags["atmospheric_loss"].get() else 0,
                "scintillation_loss": self.input_handler.get_numeric_value("scintillation_loss") if self.input_handler.flags["scintillation_loss"].get() else 0,
                "polarization_loss": self.input_handler.get_numeric_value("polarization_loss") if self.input_handler.flags["polarization_loss"].get() else 0,
                
                "rain_rate": self.input_handler.get_numeric_value("rain_rate") if self.input_handler.flags["rain_rate"].get() else 0,
                "link_margin": self.input_handler.get_numeric_value("link_margin") if self.input_handler.flags["link_margin"].get() else 0,
                "beam_edge_loss": self.input_handler.get_numeric_value("beam_edge_loss") if self.input_handler.flags["beam_edge_loss"].get() else 0,
                "scan_loss": self.input_handler.get_numeric_value("scan_loss") if self.input_handler.flags["scan_loss"].get() else 0,
            
                "interference_psd": self.input_handler.get_numeric_value("interference_psd") if self.input_handler.flags["interference_psd"].get() else -math.inf,  # 新增干扰参数

                # "distance": self.input_handler.get_numeric_value("distance") if self.input_handler.flags["distance"].get() else 0,  # km, 收发端2d距离 km
            }

            # 获取收发端参数
            if self.link_type_var.get() == "星-地上行": # 星-地上行链路
                # 发端参数：终端
                input_params["tx_eirp"] = self.input_handler.get_numeric_value("terminal_eirp")  # dBW
                # 收端参数：卫星
                input_params["rx_antenna_gain"] = self.input_handler.get_numeric_value("satellite_antenna_gain")  # dBi
                input_params["rx_noise_temp"] = self.input_handler.get_numeric_value("satellite_noise_temp")
                input_params["rx_noise_figure"] = self.input_handler.get_numeric_value("satellite_noise_figure")
            elif self.link_type_var.get() == "星-地下行": # 星-地下行链路
                # 发端参数：卫星
                input_params["tx_eirp"] = self.input_handler.get_numeric_value("satellite_eirp")  # dBW
                # 收端参数：终端
                input_params["rx_antenna_gain"] = self.input_handler.get_numeric_value("terminal_antenna_gain")  # dBi
                input_params["rx_noise_temp"] = self.input_handler.get_numeric_value("terminal_noise_temp")
                input_params["rx_noise_figure"] = self.input_handler.get_numeric_value("terminal_noise_figure")
        
        else: # 地对地场景
            # 获取公共输入参数
            input_params = {
                "frequency": self.input_handler.get_numeric_value("frequency"),  # GHz,
                "bandwidth": self.input_handler.get_numeric_value("bandwidth"),  # MHz,
                "distance": self.input_handler.get_numeric_value("distance"),  # km, 收发端2d距离 km
                "beam_edge_loss": self.input_handler.get_numeric_value("beam_edge_loss") if self.input_handler.flags["beam_edge_loss"].get() else 0,
                "interference_psd": self.input_handler.get_numeric_value("interference_psd") if self.input_handler.flags["interference_psd"].get() else -math.inf,  # 新增干扰参数
            }
            # 合并进地面信道状态信息 
            input_params.update(self.input_handler.get_terrestrial_link_parameters()) 
            print(input_params)
            # 获取收发端参数
            if self.link_type_var.get() == "地-地上行": # 地-地上行链路
                # 发端参数：终端
                input_params["tx_eirp"] = self.input_handler.get_numeric_value("terminal_eirp")  # dBW
                # 收端参数：基站
                input_params["rx_antenna_gain"] = self.input_handler.get_numeric_value("bs_antenna_gain")  # dBi
                input_params["rx_noise_temp"] = self.input_handler.get_numeric_value("bs_noise_temp")
                input_params["rx_noise_figure"] = self.input_handler.get_numeric_value("bs_noise_figure")
            elif self.link_type_var.get() == "地-地下行": # 星-地下行链路
                # 发端参数：基站
                input_params["tx_eirp"] = self.input_handler.get_numeric_value("bs_eirp")  # dBW
                # 收端参数：终端
                input_params["rx_antenna_gain"] = self.input_handler.get_numeric_value("terminal_antenna_gain")  # dBi
                input_params["rx_noise_temp"] = self.input_handler.get_numeric_value("terminal_noise_temp")
                input_params["rx_noise_figure"] = self.input_handler.get_numeric_value("terminal_noise_figure")

        return input_params

    def calculate(self):
        try:
            self.status_var.set("正在计算...")
            input_params = self._get_input_params()

            calculator = LinkCalculator()
            if self.link_type_var.get() == "星-地上行" or self.link_type_var.get() == "星-地下行":
                # 执行计算 
                self.results_temp = calculator.perform_calculations_sat(input_params)  # 将结果保存为类属性

                results = {
                    "链路状态": [
                        ("（位于波束中心的）终端仰角", self.results_temp["terminal_elevation_angle"], "°"),  # 终端仰角作为输出参数
                        ("星地距离", self.results_temp["distance"], "km"),
                        ("路径损耗", self.results_temp["path_loss"], "dB"),
                        ("雨衰", self.results_temp["rain_fade"], "dB"),
                    ],
                    "链路性能": [
                        ("接收信号功率谱密度", self.results_temp["received_signal_psd"], "dBm/MHz"),
                        ("噪声功率谱密度", self.results_temp["noise_psd"], "dBm/MHz"),
                        ("C/N", self.results_temp["c_to_n"], "dB"),
                        ("C/(N+I)", self.results_temp["c_to_n_plus_i"], "dB"),  # 新增C/(N+I)
                    ],
                }
            else: # 地对地场景
                # 执行计算 
                self.results_temp = calculator.perform_calculations_terrestrial(input_params)  # 将结果保存为类属性

                results = {
                    "链路状态": [
                        ("路径损耗", self.results_temp["path_loss"], "dB"),
                    ],
                    "链路性能": [
                        ("接收信号功率谱密度", self.results_temp["received_signal_psd"], "dBm/MHz"),
                        ("噪声功率谱密度", self.results_temp["noise_psd"], "dBm/MHz"),
                        ("C/N", self.results_temp["c_to_n"], "dB"),
                        ("C/(N+I)", self.results_temp["c_to_n_plus_i"], "dB"),  # 新增C/(N+I)
                    ],
                }

            # 更新G/T值显示框
            if self.link_type_var.get() == "星-地上行" and self.satellite_gt_label is not None:
                self.satellite_gt_label.configure(text=format_result(self.results_temp["gt_ratio"]))
            elif self.link_type_var.get() in ["星-地下行", "地-地下行"] and self.terminal_gt_label is not None:
                self.terminal_gt_label.configure(text=format_result(self.results_temp["gt_ratio"]))
            elif self.link_type_var.get() =="地-地上行" and self.bs_gt_label is not None:
                self.bs_gt_label.configure(text=format_result(self.results_temp["gt_ratio"]))

            if self.link_type_var.get() == "星-地上行": # 星-地上行链路
                results["链路性能"].append(("卫星G/T值", self.results_temp["gt_ratio"], "dB/K"))
            elif self.link_type_var.get() in ["星-地下行", "地-地下行"]: # 下行链路
                results["链路性能"].append(("终端G/T值", self.results_temp["gt_ratio"], "dB/K"))
            elif self.link_type_var.get() =="地-地上行": # 地-地上行链路
                results["链路性能"].append(("基站G/T值", self.results_temp["gt_ratio"], "dB/K"))

            self.result_display.update_results(results, self.link_type_var.get())
            self.status_var.set("计算完成")
        except Exception as e:
            messagebox.showerror("计算错误", f"计算过程中出现错误: {str(e)}")
            self.status_var.set("计算失败，请检查输入")


    def reset(self):
        self.input_handler.reset_params()
        self.status_var.set("支持公式化输入，例如：sin(30)、arctan(1)、53-10*log(16)")

    def generate_report(self):
        try:
            # 获取当前时间作为文件名的一部分
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            # 弹出文件保存对话框，让用户选择保存路径
            file_path = filedialog.asksaveasfilename(
                defaultextension=".xlsx",
                filetypes=[("Excel 文件", "*.xlsx")],
                initialfile=f"卫星链路预算仿真报告_{timestamp}.xlsx"
            )
            if not file_path:  # 如果用户取消选择，直接返回
                return

            # 创建一个新的 Excel 工作簿
            wb = Workbook()
            ws = wb.active
            ws.title = "仿真报告"

            # 设置表格样式
            header_font = Font(name='微软雅黑', size=12, bold=True)
            data_font = Font(name='微软雅黑', size=11)
            border = Border(left=Side(style='thin'), 
                         right=Side(style='thin'), 
                         top=Side(style='thin'), 
                         bottom=Side(style='thin'))
            alignment = Alignment(horizontal='center', vertical='center')

            # 定义参数标签
            param_labels = {
                "frequency": ("频率", "GHz"),
                "bandwidth": ("带宽", "MHz"),
                "satellite_height": ("卫星高度", "km"),
                "satellite_scan_angle": ("卫星扫描角", "°"),
                "terminal_eirp": ("终端EIRP", "dBW"),
                "satellite_antenna_gain": ("卫星天线增益", "dBi"),
                "satellite_noise_figure": ("卫星噪声系数", "dB"),
                "satellite_noise_temp": ("卫星天线噪声温度", "K"),
                "satellite_eirp": ("卫星EIRP", "dBW"),
                "terminal_noise_figure": ("终端噪声系数", "dB"),
                "terminal_noise_temp": ("终端天线噪声温度", "K"),
                "terminal_antenna_gain": ("终端天线增益", "dBi"),
                "atmospheric_loss": ("大气损耗", "dB"),
                "scintillation_loss": ("闪烁损耗", "dB"),
                "polarization_loss": ("极化损耗", "dB"),
                "beam_edge_loss": ("波束边缘损耗", "dB"),
                "scan_loss": ("扫描损耗", "dB"),
                "rain_rate": ("降雨率", "mm/h"),
                "link_margin": ("链路余量", "dB"),
                "interference_psd": ("干扰信号功率谱密度", "dBm/MHz")
            }

            # 写入报告标题
            ws.merge_cells('A1:C1')
            title_cell = ws['A1']
            title_cell.value = "卫星链路预算仿真报告"
            title_cell.font = Font(name='微软雅黑', size=14, bold=True)
            title_cell.alignment = alignment
            title_cell.border = border

            ws.append([])  # 空行

            # 写入输入参数
            ws.append(["输入参数", "单位", "值"])
            ws['A'+str(ws.max_row)].font = header_font
            ws['B'+str(ws.max_row)].font = header_font
            ws['C'+str(ws.max_row)].font = header_font

            # 写入输入参数并应用样式
            for param, (label, unit) in param_labels.items():
                if param in self.input_handler.params:
                    flag_key = {
                        "atmospheric_loss": "atmospheric_loss",
                        "scintillation_loss": "scintillation_loss",
                        "polarization_loss": "polarization_loss",
                        "rain_rate": "rain_rate",
                        "beam_edge_loss": "beam_edge_loss",
                        "scan_loss": "scan_loss",
                        "link_margin": "link_margin",
                        "interference_psd": "interference_psd"
                    }.get(param, None)

                    # 如果参数没有标志位，或者标志位为 True，则显示
                    if not flag_key or self.input_handler.flags[flag_key].get():
                        value = self.input_handler.params[param].get()
                        try:
                            numeric_value = float(safe_eval(value)) if value else 0
                            ws.append([label, unit, numeric_value])
                        except:
                            ws.append([label, unit, 0])

            # 设置列宽
            ws.column_dimensions['A'].width = 30
            ws.column_dimensions['B'].width = 15
            ws.column_dimensions['C'].width = 20

            # 写入计算结果
            if hasattr(self, "results_temp"):
                ws.append([])  # 空行
                ws.append(["计算结果", "单位", "值"])
                ws['A'+str(ws.max_row)].font = header_font
                ws['B'+str(ws.max_row)].font = header_font
                ws['C'+str(ws.max_row)].font = header_font

                # 定义结果项与results_temp中键的映射关系
                result_mapping = {
                    "（位于波束中心的）终端仰角": "terminal_elevation_angle",
                    "星地距离": "distance",
                    "距离": "distance",
                    "路径损耗": "path_loss",
                    "雨衰": "rain_fade",
                    "接收信号功率谱密度": "received_signal_psd",
                    "噪声功率谱密度": "noise_psd",
                    "C/N": "c_to_n",
                    "C/(N+I)": "c_to_n_plus_i",
                    "卫星G/T值": "gt_ratio",
                    "终端G/T值": "gt_ratio",
                    "基站G/T值": "gt_ratio"
                }

                # 定义结果项
                if self.link_type_var.get() == "星-地上行" or self.link_type_var.get() == "星-地下行":
                    result_items = [
                        ("（位于波束中心的）终端仰角", "度"),
                        ("星地距离", "km"),
                        ("路径损耗", "dB"),
                        ("雨衰", "dB"),
                        ("接收信号功率谱密度", "dBm/MHz"),
                        ("噪声功率谱密度", "dBm/MHz"),
                        ("C/N", "dB"),
                        ("C/(N+I)", "dB"),
                    ]
                else: # 地对地场景
                    result_items = [
                        ("距离", "km"),
                        ("路径损耗", "dB"),
                        ("接收信号功率谱密度", "dBm/MHz"),
                        ("噪声功率谱密度", "dBm/MHz"),
                        ("C/N", "dB"),
                        ("C/(N+I)", "dB"),
                    ]

                # 根据链路类型添加G/T值
                if self.link_type_var.get() == "星-地上行":
                    result_items.append(("卫星G/T值", "dB/K"))
                elif self.link_type_var.get() == "星-地下行" or self.link_type_var.get() == "地-地下行":
                    result_items.append(("终端G/T值", "dB/K"))
                else:  # 地对地场景
                    result_items.append(("基站G/T值", "dB/K"))

                # 写入结果并应用样式
                for label, unit in result_items:
                    # 通过映射关系获取正确的键
                    key = result_mapping[label]
                    # 确保获取的值是数值类型
                    try:
                        value = float(self.results_temp[key])
                        ws.append([label, unit, value])
                    except:
                        ws.append([label, unit, 0])

            # 应用边框样式到所有单元格
            for row in ws.iter_rows():
                for cell in row:
                    cell.border = border

            # 保存 Excel 文件
            wb.save(file_path)
            messagebox.showinfo("报告生成成功", f"仿真报告已保存至:\n{file_path}")
        except Exception as e:
            messagebox.showerror("报告生成失败", f"生成仿真报告时出错:\n{str(e)}")

    def show_detailed_calculation(self):
        """显示详细计算步骤"""
        try:
            # 获取输入参数
            input_params = self._get_input_params()
            # 调用详细计算方法
            calculator = LinkCalculator()
            details = calculator.detailed_calculation(input_params)
            
            # 创建新窗口显示结果
            detail_window = ctk.CTkToplevel(self.root)
            detail_window.title("详细计算步骤")
            detail_window.geometry("800x600")

            # 将新窗口设置为主窗口的子窗口
            detail_window.transient(self.root)
            # 将新窗口提升到最前面
            detail_window.lift()

            # 创建文本框显示详细步骤
            text_box = ctk.CTkTextbox(detail_window, wrap="word")
            text_box.pack(fill="both", expand=True, padx=10, pady=10)
            
            # 格式化输出
            for detail in details:
                text_box.insert("end", f"步骤: {detail['步骤']}\n")
                text_box.insert("end", f"公式: {detail['公式']}\n")
                text_box.insert("end", f"参数: {detail['参数']}\n")
                text_box.insert("end", f"结果: {detail['结果']}\n\n")
                
        except Exception as e:
            messagebox.showerror("错误", f"无法显示详细计算步骤: {str(e)}")

    def toggle_theme(self):
        """ 深色模式 """
        ctk.set_appearance_mode("dark" if self.theme_switch.get() else "light")

    # 添加新的方法实现单位转换器
    def show_unit_converter(self):
        """显示单位转换器窗口"""
        converter_window = ctk.CTkToplevel(self.root)
        converter_window.title("单位转换器")
        converter_window.geometry("600x400")
        
        # 主框架
        main_frame = ctk.CTkFrame(converter_window)
        main_frame.pack(fill="both", expand=True, padx=10, pady=10)

        # 创建转换器框架
        converter_frame = ctk.CTkFrame(main_frame, fg_color="transparent")
        converter_frame.pack(fill="both", expand=True)

        # 定义转换类型
        converters = [
            ("dBW ↔ dBm", "dBW", "dBm", lambda x: x + 30, lambda x: x - 30),
            ("W ↔ dBW", "W", "dBW", lambda x: 10 * math.log10(x), lambda x: 10 ** (x / 10)),
            ("dBm/MHz ↔ dBm/RE", "dBm/MHz", "dBm/RE", 
            lambda x: x - 10 * math.log10(1000/15), 
            lambda x: x + 10 * math.log10(1000/15)),
            ("KHz ↔ MHz", "KHz", "MHz", lambda x: x / 1000, lambda x: x * 1000),
        ]

        # 创建每个转换器的输入框
        for i, (title, unit1, unit2, func1, func2) in enumerate(converters):
            frame = ctk.CTkFrame(converter_frame, fg_color="transparent")
            frame.pack(fill="x", pady=5)

            # 标题
            ctk.CTkLabel(frame, text=title, font=("微软雅黑", 12, "bold")).pack(side=tk.TOP, pady=5)

            # 输入框容器
            input_container = ctk.CTkFrame(frame, fg_color="transparent")
            input_container.pack(fill="x", pady=2)

            # 输入框1
            input_frame1 = ctk.CTkFrame(input_container, fg_color="transparent")
            input_frame1.pack(side=tk.LEFT, fill="x", expand=True, padx=5)
            ctk.CTkLabel(input_frame1, text=unit1, width=80).pack(side=tk.LEFT)
            entry1 = ctk.CTkEntry(input_frame1)
            entry1.pack(side=tk.RIGHT, fill="x", expand=True)

            # 输入框2
            input_frame2 = ctk.CTkFrame(input_container, fg_color="transparent")
            input_frame2.pack(side=tk.RIGHT, fill="x", expand=True, padx=5)
            ctk.CTkLabel(input_frame2, text=unit2, width=80).pack(side=tk.LEFT)
            entry2 = ctk.CTkEntry(input_frame2)
            entry2.pack(side=tk.RIGHT, fill="x", expand=True)

            # 绑定事件
            entry1.bind("<FocusIn>", lambda e, e1=entry1: self._on_converter_focus_in(e1))
            entry1.bind("<FocusOut>", lambda e, e1=entry1, e2=entry2, f=func1: 
                        self._on_converter_focus_out(e1, e2, f))
            entry2.bind("<FocusIn>", lambda e, e2=entry2: self._on_converter_focus_in(e2))
            entry2.bind("<FocusOut>", lambda e, e1=entry1, e2=entry2, f=func2: 
                        self._on_converter_focus_out(e2, e1, f))

    def _on_converter_focus_in(self, entry):
        """当输入框获得焦点时，显示原始公式"""
        if hasattr(entry, 'raw_formula'):
            entry.delete(0, tk.END)
            entry.insert(0, entry.raw_formula)

    def _on_converter_focus_out(self, src_entry, dst_entry, convert_func):
        """当输入框失去焦点时，计算并显示结果"""
        expr = src_entry.get().strip()
        src_entry.raw_formula = expr
        try:
            value = safe_eval(expr)
            if value is not None:
                result = convert_func(float(value))
                dst_entry.delete(0, tk.END)
                dst_entry.insert(0, f"{result:.4f}")
        except Exception as e:
            dst_entry.delete(0, tk.END)


    def _update_converter(self, src_entry, dst_entry, convert_func):
        """更新转换结果"""
        try:
            value_str = src_entry.get().strip()
            if not value_str:
                dst_entry.delete(0, tk.END)
                return
            value = float(value_str)
            result = convert_func(value)
            dst_entry.delete(0, tk.END)
            dst_entry.insert(0, f"{result:.4f}")
        except ValueError:
            dst_entry.delete(0, tk.END)


if __name__ == "__main__":
    root = ctk.CTk()
    app = SatelliteLinkBudgetCalculator(root)
    root.mainloop()
